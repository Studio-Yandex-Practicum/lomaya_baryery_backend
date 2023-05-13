import enum
from dataclasses import astuple
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from src.core.db.DTO_models import TasksAnalyticReportDto
from src.excel_generator.builder import AnalyticReportBuilder
from src.excel_generator.task_builder import BaseAnalyticReportSettings


class ShiftAnalyticReportSettings(BaseAnalyticReportSettings):
    """Конфигурация отчёта для выбранной смены."""

    sheet_name: str = "Отчёт по смене"
    header_data: tuple[str] = (
        "№ Задачи",
        "Задача",
        "Принята с 1-й попытки",
        "Принята с 2-й попытки",
        "Принята с 3-й попытки",
        "Кол-во пропусков задания",
        "Кол-во одобренных отчётов",
        "Кол-во отклонённых отчётов",
    )
    row_count: int = 0

    @classmethod
    @property
    def footer_data(cls):
        return (
            "ИТОГО:",
            "",
            f"=SUM(C2:C{cls.row_count})",
            f"=SUM(D2:D{cls.row_count})",
            f"=SUM(E2:E{cls.row_count})",
            f"=SUM(F2:F{cls.row_count})",
            f"=SUM(G2:G{cls.row_count})",
            f"=SUM(H2:H{cls.row_count})",
        )


class AnalyticShiftReportBuilder(AnalyticReportBuilder):
    async def generate_report(
        self,
        description: str,
        data: tuple[TasksAnalyticReportDto],
        workbook: Workbook,
        analytic_task_report_full: BaseAnalyticReportSettings,
    ) -> Workbook:
        """Генерация листа с данными."""
        worksheet = self._create_sheet(workbook, sheet_name=analytic_task_report_full.sheet_name)
        analytic_task_report_full.row_count = 0
        self.__add_description(worksheet, description, analytic_task_report_full)
        self.__add_header(worksheet, analytic_task_report_full)
        self.__add_data(worksheet, data, analytic_task_report_full)
        self.__add_footer(worksheet, analytic_task_report_full)
        self.__apply_styles(worksheet)
        return workbook

    def __add_row(
        self,
        worksheet: Worksheet,
        analytic_task_report: BaseAnalyticReportSettings,
        data: tuple[str | int],
    ) -> None:
        analytic_task_report.row_count += 1
        for index, value in enumerate(data, start=1):
            worksheet.cell(row=analytic_task_report.row_count, column=index, value=value)

    @staticmethod
    async def get_report_response(workbook: Workbook) -> BytesIO:
        """Создание ответа."""
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    @staticmethod
    def create_workbook() -> Workbook:
        """Генерация excel файла."""
        workbook = Workbook()
        workbook.remove_sheet(workbook.active)
        return workbook

    @staticmethod
    def _create_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
        """Создаёт лист внутри отчёта."""
        return workbook.create_sheet(sheet_name)

    def __add_description(
        self,
        worksheet: Worksheet,
        description: str,
        analytic_task_report: BaseAnalyticReportSettings,
    ) -> None:
        """Заполняет описание отчета."""
        self.__add_row(worksheet, analytic_task_report, (description,))

    def __add_header(self, worksheet: Worksheet, analytic_task_report: BaseAnalyticReportSettings) -> None:
        """Заполняет первые строки в листе."""
        self.__add_row(worksheet, analytic_task_report, analytic_task_report.header_data)

    def __add_data(
        self,
        worksheet: Worksheet,
        data: tuple[TasksAnalyticReportDto],
        analytic_task_report: BaseAnalyticReportSettings,
    ) -> None:
        """Заполняет строки данными из БД."""
        for task in data:
            self.__add_row(worksheet, analytic_task_report, data=astuple(task))

    def __add_footer(self, worksheet: Worksheet, analytic_task_report: BaseAnalyticReportSettings) -> None:
        """Заполняет последнюю строку в листе."""
        self.__add_row(worksheet, analytic_task_report, data=analytic_task_report.footer_data)

    def __apply_styles(self, worksheet: Worksheet):
        """Задаёт форматирование отчёта."""
        # задаём стиль для хэдера
        rows = list(worksheet.iter_rows())
        header_cells = rows[1]
        for cell in header_cells:
            cell.font = self.Styles.FONT_BOLD.value
            cell.alignment = self.Styles.ALIGNMENT_HEADER.value
        # задаём стиль для ячеек с данными
        data_rows = rows[2:-1]
        for row in data_rows:
            for cell in row:
                cell.font = self.Styles.FONT_STANDART.value
                cell.alignment = self.Styles.ALIGNMENT_STANDART.value
        # задаём стиль для футера
        footer_cells = rows[-1]
        for cell in footer_cells:
            cell.font = self.Styles.FONT_BOLD.value
            cell.alignment = self.Styles.ALIGNMENT_STANDART.value
        # задаём стиль границ и колонок
        for row in rows:
            for cell in row:
                cell.border = self.Styles.BORDER.value
        worksheet.column_dimensions["B"].width = self.Styles.WIDTH.value

    class Styles(enum.Enum):
        FONT_BOLD = Font(name='Times New Roman', size=11, bold=True)
        FONT_STANDART = Font(name='Times New Roman', size=11, bold=False)
        ALIGNMENT_HEADER = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ALIGNMENT_STANDART = Alignment(horizontal='left', vertical='center', wrap_text=True)
        BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
        )
        WIDTH = 30
