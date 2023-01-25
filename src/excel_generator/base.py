import enum
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet


class ExcelBaseGenerator:
    """Базовый генератор excel отчётов."""

    def __init__(self) -> None:
        self.workbook = Workbook()
        self.workbook.remove_sheet(self.workbook.active)

    def __clear_workbook(self):
        """Очищает Workbook после сохранения данных."""
        sheets = self.workbook.get_sheet_names()
        for name in sheets:
            sheet_to_remove = self.workbook.get_sheet_by_name(name)
            self.workbook.remove_sheet(sheet_to_remove)

    async def save_report_to_stream(self) -> BytesIO:
        """Сохраняет отчёт в буфер памяти."""
        stream = BytesIO()
        self.workbook.save(stream)
        stream.seek(0)
        self.__clear_workbook()
        return stream

    def fill_row(self, sheet: Worksheet, data_list: tuple[str], row_number: int) -> None:
        """Заполняет строку данными."""
        for idx, value in enumerate(data_list):
            sheet.cell(row=row_number, column=idx + 1, value=value)

    def make_styling(self, sheet: Worksheet, max_row: int):
        """Задаёт форматирование отчёта."""
        # задаём стиль для хэдера
        header_cells = list(sheet.iter_rows())[0]
        for cell in header_cells:
            cell.font = self.Styles.FONT_BOLD.value
            cell.alignment = self.Styles.ALIGNMENT_HEADER.value
        # задаём стиль для ячеек с данными
        data_rows = list(sheet.iter_rows())[1:max_row - 1]  # fmt: skip
        for row in data_rows:
            for cell in row:
                cell.font = self.Styles.FONT_STANDART.value
                cell.alignment = self.Styles.ALIGNMENT_STANDART.value
        # задаём стиль для футера
        footer_cells = list(sheet.iter_rows())[max_row - 1]
        for cell in footer_cells:
            cell.font = self.Styles.FONT_BOLD.value
            cell.alignment = self.Styles.ALIGNMENT_STANDART.value
        # задаём стиль границ и колонок
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = self.Styles.BORDER.value
        sheet.column_dimensions["A"].width = self.Styles.WIDTH.value

    class Sheets(str, enum.Enum):
        TASKS = "Задачи"

    class Styles(enum.Enum):
        FONT_BOLD = Font(name='Times New Roman', size=11, bold=True)
        FONT_STANDART = Font(name='Times New Roman', size=11, bold=False)
        ALIGNMENT_HEADER = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ALIGNMENT_STANDART = Alignment(horizontal='left', vertical='center', wrap_text=True)
        BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
        )
        WIDTH = 50
