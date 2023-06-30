import dataclasses
import enum
from dataclasses import astuple
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from src.core.db.DTO_models import TasksAnalyticReportDto
from src.excel_generator.base_analytic_report_settings import BaseAnalyticReportSettings


class AnalyticReportBuilder:
    """Интерфейс строителя."""

    def add_sheet(
        self,
        description: str,
        data: tuple[TasksAnalyticReportDto],
        workbook: Workbook,
        analytic_report_settings: BaseAnalyticReportSettings,
    ) -> Workbook:
        """Генерация листа с данными."""
        worksheet = self._create_sheet(workbook, sheet_name=analytic_report_settings.sheet_name)
        analytic_report_settings.row_count = 0
        self.__add_description(worksheet, description, analytic_report_settings)
        self.__add_header(worksheet, analytic_report_settings)
        self.__add_data(worksheet, data, analytic_report_settings)
        self.__add_footer(worksheet, analytic_report_settings)
        self.__apply_styles(worksheet)
        return workbook

    def __add_row(
        self,
        worksheet: Worksheet,
        analytic_report_settings: BaseAnalyticReportSettings,
        data: tuple[str | int],
    ) -> None:
        analytic_report_settings.row_count += 1
        for index, value in enumerate(data, start=1):
            worksheet.cell(row=analytic_report_settings.row_count, column=index, value=value)

    @staticmethod
    def get_report_response(workbook: Workbook) -> BytesIO:
        """Создание ответа."""
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    @staticmethod
    def create_workbook() -> Workbook:
        """Генерация excel файла."""
        workbook = Workbook()
        workbook.remove_sheet(workbook.active)
        return workbook

    @staticmethod
    def _create_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
        """Создаёт лист внутри отчёта."""
        return workbook.create_sheet(sheet_name)

    def __add_description(
        self,
        worksheet: Worksheet,
        description: str,
        analytic_report_settings: BaseAnalyticReportSettings,
    ) -> None:
        """Заполняет описание отчета."""
        self.__add_row(worksheet, analytic_report_settings, (description,))

    def __add_header(self, worksheet: Worksheet, analytic_report_settings: BaseAnalyticReportSettings) -> None:
        """Заполняет первые строки в листе."""
        self.__add_row(worksheet, analytic_report_settings, analytic_report_settings.header_data)

    def __add_data(
        self,
        worksheet: Worksheet,
        data: tuple[TasksAnalyticReportDto],
        analytic_report_settings: BaseAnalyticReportSettings,
    ) -> None:
        """Заполняет строки данными из БД."""
        for task in data:
            if dataclasses.is_dataclass(task):
                task = astuple(task)
            self.__add_row(worksheet, analytic_report_settings, task)

    def __add_footer(self, worksheet: Worksheet, analytic_report_settings: BaseAnalyticReportSettings) -> None:
        """Заполняет последнюю строку в листе."""
        self.__add_row(worksheet, analytic_report_settings, data=analytic_report_settings.footer_data)

    def __apply_styles(self, worksheet: Worksheet):
        """Задаёт форматирование отчёта."""
        rows = list(worksheet.iter_rows())
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(rows[0]))
        # задаём стиль для описания отчета
        description_rows = rows[0]
        for cell in description_rows:
            cell.font = self.Styles.FONT_STANDART.value
            cell.alignment = self.Styles.DESCRIPTION_ALIGNMENT.value
        # задаём стиль для хэдера
        header_cells = rows[1]
        for cell in header_cells:
            cell.font = self.Styles.FONT_BOLD.value
            cell.alignment = self.Styles.ALIGNMENT_HEADER.value
        # задаём стиль для ячеек с данными
        data_rows = rows[2:-1]
        for row in data_rows:
            for cell in row:
                cell.font = self.Styles.FONT_STANDART.value
                cell.alignment = self.Styles.ALIGNMENT_STANDART.value
        # задаём стиль для футера
        footer_cells = rows[-1]
        for cell in footer_cells:
            cell.font = self.Styles.FONT_BOLD.value
            cell.alignment = self.Styles.ALIGNMENT_STANDART.value
        # задаём стиль границ и колонок
        for row in rows:
            for cell in row:
                cell.border = self.Styles.BORDER.value
        worksheet.column_dimensions["B"].width = self.Styles.WIDTH.value
        worksheet.row_dimensions[1].height = self.Styles.HEIGHT.value

    class Styles(enum.Enum):
        FONT_BOLD = Font(name='Times New Roman', size=11, bold=True)
        FONT_STANDART = Font(name='Times New Roman', size=11, bold=False)
        ALIGNMENT_HEADER = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ALIGNMENT_STANDART = Alignment(horizontal='left', vertical='center', wrap_text=True)
        DESCRIPTION_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrap_text=True, shrink_to_fit=True)
        BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
        )
        WIDTH = 50
        HEIGHT = 55
