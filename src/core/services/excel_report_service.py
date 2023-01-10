import enum
from datetime import datetime
from io import BytesIO

from fastapi import Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.core.db.repository.task_repository import TaskRepository
from src.core.settings import settings


class ExcelReportService:
    """Сервис для создания excel отчётов."""

    def __init__(self, task_repository: TaskRepository = Depends()) -> None:
        self.__task_repository = task_repository

    def __get_report_template(self) -> Workbook:
        """Возвращает excel шаблон."""
        workbook = load_workbook(settings.report_template_path)
        workbook.template = False
        return workbook

    def __delete_unused_lists(self, workbook: Workbook, reports: tuple) -> None:
        """Удаляет неиспользуемые листы."""
        for name in workbook.sheetnames:
            if name not in reports:
                workbook.remove(workbook[name])

    def __save_report_to_stream(self, workbook: Workbook) -> BytesIO:
        """Сохраняет отчёт в буфер памяти."""
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    async def generate_report(self, reports: tuple[str]) -> StreamingResponse:
        """Создаёт excel отчёт."""
        workbook = self.__get_report_template()
        # создание отчётов
        if self.Sheets.TASKS.value in reports:
            await self.__create_tasks_statistics_report(workbook)
        if self.Sheets.TEST.value in reports:
            await self.__create_test_report(workbook)
        # удаление неиспользуемых листов
        self.__delete_unused_lists(workbook, reports)
        stream = self.__save_report_to_stream(workbook)
        filename = f"report_{datetime.now()}.xlsx"
        headers = {'Content-Disposition': f'attachment; filename={filename}'}
        return StreamingResponse(stream, headers=headers)

    async def __create_tasks_statistics_report(self, workbook: Workbook) -> Worksheet:
        tasks_report = await self.__task_repository.get_tasks_statistics_report()
        sheet = workbook[self.Sheets.TASKS]
        for idx, task in enumerate(tasks_report):
            row_num = idx + 2
            sheet.cell(row=row_num, column=1, value=task.description)
            sheet.cell(row=row_num, column=2, value=task.approved)
            sheet.cell(row=row_num, column=3, value=task.declined)
            sheet.cell(row=row_num, column=4, value=task.waiting)
        return sheet

    # TODO: удалить перед мержем в develop
    async def __create_test_report(self, workbook: Workbook) -> Worksheet:
        sheet = workbook[self.Sheets.TEST]
        sheet.cell(row=1, column=1, value="тест")
        return sheet

    class Sheets(str, enum.Enum):
        TASKS = "Задачи"
        TEST = "Тест"
